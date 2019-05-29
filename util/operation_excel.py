'''excel的读写，用于读取接口用例和写入接口请求响应值'''
import  xlrd,os
from openpyxl import load_workbook
BASE_PATH = os.path.abspath(os.path.dirname(os.path.abspath(__file__)) + '\..')
# excel_path=BASE_PATH+r'\config\interface_testcases.xlsx'
excel_path=BASE_PATH+r'\config\test.xlsx'

class Excel:
    def __init__(self):
        self.filename=excel_path

    def read_excel(self):
        excel=xlrd.open_workbook(self.filename)
        '''获取excel多个sheet的数据'''
        sheet_count=len(excel.sheets())
        caseInfo = []
        sheets_row=[]
        for s in range(sheet_count):
            sheet=excel.sheet_by_index(s)
            rows=sheet.nrows
            sheets_row.append(rows-1)
            cols=sheet.ncols
            headers=sheet.row_values(0)
            # caseInfo=[]
            for i in range(1,rows):
                case_info = {}
                for j in range(0,cols):
                    case_info[headers[j]]=sheet.row_values(i)[j]
                caseInfo.append(case_info)
        return caseInfo,sheets_row

    def write_respone(self,sheetIndex,row,respone):
        wb = load_workbook(self.filename)
        sheets_names = wb.sheetnames
        wb1 = wb[sheets_names[sheetIndex]]
        #wb1 = wb.active
        wb1.cell(row,10,respone)
        wb.save(self.filename)

    def write_result(self,sheetIndex,row,result):
        wb = load_workbook(self.filename)
        sheets_names=wb.sheetnames
        wb1=wb[sheets_names[sheetIndex]]
        # wb1 = wb.active
        if result:
            wb1.cell(row,11,'pass')
        else:
            wb1.cell(row,11,'fail')
        wb.save(self.filename)

    def write_fail_message(self,sheetIndex,row,message):
        wb = load_workbook(self.filename)
        sheets_names = wb.sheetnames
        wb1 = wb[sheets_names[sheetIndex]]
        #wb1 = wb.active
        if message:
            wb1.cell(row,12,message)
        wb.save(self.filename)

    def write_number(self,row):
        wb = load_workbook(self.filename)
        wb1 = wb.active
        wb1.cell(row,1,row-1)
        wb.save(self.filename)


if __name__ == '__main__':
    Excel().read_excel()
    #Excel().write_result(0,6,'test')